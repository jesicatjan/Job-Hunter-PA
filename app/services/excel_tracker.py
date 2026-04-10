"""
Excel Application Tracker – replaces Notion.
Maintains a single Excel workbook (data/applications.xlsx) with:
  Sheet 1 – Applications  (one row per application)
  Sheet 2 – Dashboard     (pivot summary)

Users can download the file any time via /export.
"""
import logging
from datetime import date, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app import database as db

logger = logging.getLogger(__name__)

XLSX_PATH = Path("./data/applications.xlsx")
XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

HEADERS = [
    "ID", "Company", "Role", "Status", "Source",
    "Salary", "Applied Date", "Follow-up Date", "Interview Date",
    "URL", "Notes",
]

STATUS_COLORS = {
    "Applied":      "FFF9C4",   # yellow
    "Interviewed":  "C8E6C9",   # green
    "Offered":      "A5D6A7",   # strong green
    "Rejected":     "FFCDD2",   # red
    "Withdrawn":    "E0E0E0",   # grey
}

HEADER_FILL = PatternFill("solid", fgColor="1565C0")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _thin_border():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def rebuild_workbook(telegram_id: int) -> Path:
    """
    Rebuild the Excel workbook from the SQLite database for this user.
    Returns the path to the saved file.
    """
    apps = db.get_applications(telegram_id)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Applications ──────────────────────────────────────
    ws = wb.active
    ws.title = "Applications"

    # Headers
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, app in enumerate(apps, start=2):
        status = app.get("status", "Applied")
        fill_color = STATUS_COLORS.get(status, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_color)

        values = [
            app.get("id"),
            app.get("company"),
            app.get("role"),
            status,
            app.get("source"),
            app.get("salary"),
            app.get("applied_date"),
            app.get("followup_date"),
            app.get("interview_date"),
            app.get("url"),
            app.get("notes"),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Hyperlink for URL column (col 10)
        url = app.get("url")
        if url:
            link_cell = ws.cell(row=row_idx, column=10)
            link_cell.hyperlink = url
            link_cell.value = "Open Job"
            link_cell.font = Font(color="1565C0", underline="single")

    # Column widths
    col_widths = [6, 20, 25, 14, 16, 20, 14, 14, 14, 12, 40]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # ── Sheet 2: Dashboard ─────────────────────────────────────────
    ds = wb.create_sheet("Dashboard")
    ds["A1"] = "Job Hunt Dashboard"
    ds["A1"].font = Font(bold=True, size=14)

    # Status counts
    from collections import Counter
    status_counts = Counter(a.get("status", "Applied") for a in apps)
    ds["A3"] = "Status"
    ds["B3"] = "Count"
    ds["A3"].font = Font(bold=True)
    ds["B3"].font = Font(bold=True)

    row = 4
    for status, count in sorted(status_counts.items()):
        ds.cell(row=row, column=1, value=status)
        ds.cell(row=row, column=2, value=count)
        row += 1

    ds.column_dimensions["A"].width = 18
    ds.column_dimensions["B"].width = 10

    # Summary stats
    total = len(apps)
    interviewed = sum(1 for a in apps if a.get("status") in ("Interviewed", "Offered"))
    offered = sum(1 for a in apps if a.get("status") == "Offered")

    ds["D3"] = "Total Applied"
    ds["E3"] = total
    ds["D4"] = "Interviews"
    ds["E4"] = interviewed
    ds["D5"] = "Offers"
    ds["E5"] = offered
    ds["D6"] = "Interview Rate"
    ds["E6"] = f"{round(interviewed/total*100)}%" if total else "0%"

    for cell_ref in ["D3", "D4", "D5", "D6"]:
        ds[cell_ref].font = Font(bold=True)

    wb.save(XLSX_PATH)
    logger.info(f"Excel workbook saved: {XLSX_PATH}")
    return XLSX_PATH


def get_workbook_path(telegram_id: int) -> Path:
    """Rebuild and return path."""
    return rebuild_workbook(telegram_id)
